v'''
	1.	确定关键词
	•	选择感兴趣的关键词，例如“新能源车”、“环保政策”、“碳中和”等。
	•	可创建一个关键词列表。
	2.	设置时间范围
	•	在爬取时需输入时间范围，2020 年 1 月至 2023 年 12 月。
	3.	分地区爬取
	•	百度指数支持分地区查看数据，可指定地区编码（如浙江杭州市为 3301)。
	•	直辖市（如北京）可以用简化编码（如 11)。
	4.	保存数据
	•	将爬取的结果保存为 CSV 或 Excel 文件，方便后续分析。
'''
import requests
from datetime import datetime
from dateutil.relativedelta import relativedelta 
import openpyxl
import time
from openpyxl.styles import Alignment, Border, Side
from tool.countryMap import province_to_cities as p2c
from tool.countryMap import city_to_code as c2c
from tool.fakeheaders import generate_fake_headers 
import random

class baidu_crawler:
    def __init__(self, Cookie, keys, startDate, endDate):
        self.Cookie = Cookie
        self.keys = keys
        self.startDate = startDate
        self.endDate = endDate

    def get_index(self, key, year, month, leap_year, code):
        words = [[{"name": key, "wordType": 1}]]
        words = str(words).replace(" ", "").replace("'", "\"")
        days = [31,29 if leap_year else 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
        startDate = f"{year}-{month}-01"
        endDate = f"{year}-{month}-{days[month-1]}"
        # 搜索指数 api
        search_url = f'http://index.baidu.com/api/SearchApi/index?area={code}&word={words}&startDate={startDate}&endDate={endDate}'
        # 内容指数 api
        content_url = f'https://index.baidu.com/api/FeedSearchApi/getFeedIndex?area={code}&word={words}&startDate={startDate}&endDate={endDate}'
        # 请求头配置
        # headers = generate_fake_headers(Cookie)
        headers = {
            "Accept": "application/json, text/plain, */*",
            "Accept-Encoding": "gzip, deflate, br",
            "Accept-Language": "zh-CN,zh-Hans;q=0.9",
            "Cipher-Text": "1734429893153_1734493102920_MqbyWeYe9ierNUYfwLeypHdL1iUYu/9EdO+vDCC06irL+r3EjBmz9NGbb/UnVP1UiauNwkfg6UC2G9mm6nbB/2wPUy4bX/4Jflr5GMC41wYvUZSr92ITB1N/kNiNPVazIie2UAVPqHqjBpz0Vw8vOBF0rMBjM58mFlC7tIntVDu9+tWnzy6gw+9ea63P8FYF1RMRU/h+FM0s7Cph5E52sWh2psWIAbRXW6hz8AaAq8xkiBQ/Pe8HbA4XLi7fJinOXCYL+ig56avZq1RZ1vFakvXRJLS53HLbioBJybzjV8Dtp24n1hbftvTgqEHas1mWsctEPcwBcybnPF6FsquSUzHoc3mZiuzwcQiDilWN9dNb8jE9JsJ3BTF2LY4cq9f6M2iSDFgD0cokcwfMw5NrsQ==",
            "Connection": "keep-alive",
            "Cookie": Cookie,
            "Host": "index.baidu.com",
            "Referer": "https://index.baidu.com/v2/main/index.html",
            "Sec-Fetch-Dest": "empty",
            "Sec-Fetch-Mode": "cors",
            "Sec-Fetch-Site": "same-origin",
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.15 (KHTML, like Gecko) Version/17.2 Safari/605.15"
        }
        search_res = requests.get(search_url, headers=headers)
        time.sleep(random.uniform(7, 8))
        content_res = requests.get(content_url, headers=headers)
        time.sleep(random.uniform(8, 9))
        search_res_json = search_res.json()
        # print(search_res_json)
        content_res_json = content_res.json()
        if search_res_json["message"] == "bad request" or content_res_json["message"] == "bad request":
            print("抓取关键词："+key+" 失败，请检查cookie或者关键词是否存在")
        else:
            return [int(search_res_json['data']['generalRatio'][0]['all']['avg'])], [int(content_res_json['data']['index'][0]['generalRatio']['avg'])]

            
    #创建日期表格
    def create_excel(self, key):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        # 设置第一行的标题
        cell = sheet.cell(row=1, column=1)  # 设置 A1 单元格
        cell.value = '         省市\n月份'
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)  # 居中且自动换行
        
        # 设置单元格对角线
        thin = Side(border_style="thin", color="000000")
        cell.border = Border(diagonal=thin, diagonalUp=False, diagonalDown=True)

        # 调整单元格大小
        sheet.row_dimensions[1].height = 40  # 设置行高
        sheet.column_dimensions['A'].width = 15  # 设置列宽

        # 计算日期范围
        start_date = datetime(self.startDate, 1, 1)
        end_date = datetime(self.endDate, 12, 31)
    
        # 逐行填充日期
        current_date = start_date
        row = 2  # 从第二行开始
        while current_date <= end_date:
            sheet[f'A{row}'] = current_date.strftime('%Y-%m')
            current_date += relativedelta(months=1)
            row += 1
    
        # 保存 Excel 文件
        search_filename, content_filename = f'{key}-月度搜索百度指数-{startDate}-{endDate}.xlsx', f'{key}-月度内容百度指数-{startDate}-{endDate}-.xlsx'
        workbook.save(search_filename)
        workbook.save(content_filename)
        return search_filename, content_filename

    #为文件写入数据
    def write_to_excel(self, file_name, key, area, data, i):
        try:
            # 打开 Excel 文件
            workbook = openpyxl.load_workbook(file_name)
            # 获取默认的工作表
            sheet = workbook.active
            # 将名称写入第一行第 i 列
            sheet.cell(row=1, column=i, value=area)
            # 将数据写入从第二行开始的第i列
            for index, value in enumerate(data, start=2):
                sheet.cell(row=index, column=i, value=value)
            # 保存文件
            workbook.save(file_name)
            if len(data) != 0 :
                print(f"{startDate}-{endDate}-{area}-关键词-{key}-数据写入成功!有效数据共{len(data)-data.count(0)}个")
        except Exception as e:
            print(f"报错: {e}")

    def crawler(self, loc):      
        search_data = []
        content_data = []
        i = 2
        if loc == "all":
            areas = c2c
            loc = "全国"
        else:
            areas = {}
            for city in loc:
                areas[city] = c2c.get(city)
        num = 0
        for key in keys:
            search_filename, content_filename = self.create_excel(key)
            print(search_filename+"创建成功！")
            print(content_filename+"创建成功！")  
            for area, code in areas.items():
                num = num + 1
                if num % 5 == 0:
                    time.sleep(random.uniform(15, 25))
                for year in range(startDate, endDate + 1):
                    # 根据年份判断是否为闰年
                    if (year % 4 == 0 and year % 100 != 0) or (year % 400 == 0):
                        leap_year = True
                    else:
                        leap_year = False
                    for month in range(1,13):
                        print(f"正在处理{area}地区{key}{year}年{month}月数据...")
                        try:
                            search_data_temp, content_data_temp = self.get_index(key, year, month, leap_year, code)
                            search_data += search_data_temp
                            content_data += content_data_temp
                        except Exception as e:
                            print(f"处理{area}地区{key}{year}年{month}月数据时出错: {e}")
                            continue

                self.write_to_excel(search_filename, key, area, search_data, i)
                self.write_to_excel(content_filename, key, area, content_data, i)
                i = i + 1
                search_data = []
                content_data = []
            i = 2
            print(f"{key}数据收集完成！")
        print("程序运行结束！")


if __name__=="__main__":    
    # 默认关键词列表
    keys = ["冰雪大世界", "洪崖洞","五一广场", "淄博烧烤", "大唐不夜城"]
    # 固定时间
    startDate = 2020
    endDate = 2023

    Cookie = 'RT="z=1&dm=baidu.com&si=68f13be4-0d4b-4e6c-99f2-80b6afa4e1bf&ss=m4tcdlyt&sl=8&tt=b6v&bcn=https%3A%2F%2Ffclog.baidu.com%2Flog%2Fweirwood%3Ftype%3Dperf&ld=10io"; ab_sr=1.0.1_ZTVjNjA1YWM2YTY1YjMyZjAwOTg3YTUyNTIzMGY5ZTVlYzU5N2Q1NzA5MTQ2MDhhNmMwZWI3M2M3Y2VhOGExMWVkYTVkNzVkYzU3YTVjYTIzMTI5NDBhMGUyZmU2ZGU3Y2FjZTA2MzNkNjZkNjMxMTJkYjAwMjg2ZmRjNjIwZTc2ZjhhNmFmOTU1MzlhZWVlNzk3OGJjN2M5YjBmMDNjMg==; Hm_lpvt_d101ea4d2a5c67dab98251f0b5de24dc=1734493094; Hm_lvt_d101ea4d2a5c67dab98251f0b5de24dc=1733493447; bdindexid=m2sneogctm2nm5pnpk36ufim54; CPID_212=61784094; CPTK_212=637789256; __cas__id__212=61784094; __cas__rn__=484505539; __cas__st__212=6eeebf4d1bac9f246c31fd3045b27a023084387e16d4952b1ac583d9f573dbde2d8954199dd1fdeed4f205ef; SIGNIN_UC=70a2711cf1d3d9b1a82d2f87d633bd8a04845055399nBxNbxRYKiuDGRlcjHmIDsX06MKKGUn4dciP5hhniyzR5ZrsbHv%2BTXy5uAf1uNEFkJcGppTcgoguWOYbwADeFeq8qqDrxv3OUwWMU%2F3rrwnrsY2bB7brJ1fMoZfg5k%2F0V9CCh8MbBwzsr%2BX%2BFtgoaoULNIZbLNNJw7Z0RJXOo%2BM9YxwqULOoxNtQQLZyDoYtWtH%2BCiavcS4bd9OlQHMFndJY6igKZxdWysJ%2BUmvyDtel2elz4IpiVcoZjEimaDnaYUzk2AbVSzXp2kmb2AMtoCVCUajrNtArPloWI5XQLn8%3D77160794189163499742941651462977; BDUSS=XNMZVRud35Tai1-c1Y2V25UNDNTYnBwRThQcHd2MkNmdjhiNWdZLTgtMmgwSWxuRVFBQUFBJCQAAAAAAQAAAAEAAABNYwGTRm9yQ3Jhd2xlcjEAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAKFDYmehQ2JnN; ppfuid=FOCoIC3q5fKa8fgJnwzbE0LGziLN3VHbX8wfShDP6RCsfXQp/69CStRUAcn/QmhIlFDxPrAc/s5tJmCocrihdwitHd04Lvs3Nfz26Zt2honnQBqGZJ1+FooMMOy9VfBMHd4KSE2lo7q2jM/4leBJXqt+nMkIsYwpNA4XBVB5BGpKv8rf9RxKxeW8+CymWcpQa9KC4CVi2o94BMnih78wHfKlxopkhSdy8TWv4iu/xwHqAylyX9RqpcWHFB7PbPfTsAWTof8FuLBgffZ/0Uxl9589krRDcZw2i5aBqx+5lkk1v2iwj13daM+9aWJ5GJCQu/SUbF5jV5AUyz/jBiIgKdxDT28FggPkjANiNTdVH47OQHvyhUQrrHDS6OPGdMUXeHCBgi0CWDVWKfMtqsX5XhIgrMrxFpGm9Uos0IaIjPEvq+mAsZ7sZN6HSLYWCMfeSecr9yhMSRLVoFktEC1isIrxxJIs7iP3yZ54sBOhPbkFMb7+JFWxNGoA0JNiv6hCb0gkXpkEpISi6tVHh+hsQifjACGGz0MbLI9AAutvQNmLovQE8DrrUkOPSWZkiBwIUvxonSGS2lgiNZBxgK/Nad6P3sfvyvYhyXNwxm6SzH+Oja1l6cy9uoP7y446ILa1CLEOaV1jDkGoksNhRtn7B1VPovN1TRU04qLrmECuDGMBVR4vlhy8DqZQ1/LUEQ9mjyqP/SnZsRdyLAjuA3ESTcrCSmS6iWcmxBDT8gjuTbf5rG4+h0gsZ2eMGgzIHtuS9XnVf1zXCrpY5cuXfJTPiRwbFTm9YEOZ3QiEhlRF3TDPQd9g/PNx1HL8zvwSIj9O8YClTpPkuiWM0gZm41VCO+vNeIKxJyJ0hcHJ83oQN1+3jtOOi8LxWmDSZzbPJDJU9Bq7zt2A9A8E851l8QtBoQFIuWEGY3DMQGzE4fLtBnD2IBA1xgIrbF95h/aKYBNVXdvBhoLwXhcnXaiqXEpcvFQlonIv85FfaVbfEoKujQX2IBA1xgIrbF95h/aKYBNVh6Y0NjEKZ13xldTgKDiG2QRBJFTPsviSSEvgLGRO3YgGOv+/I3nwGp9q5hLF8/07goRUnieOy9WY3CCu1FKQrXv4Kl2tvhm/51VQHSSoTtFbHhSGlEKo+S0ciyUHoRYU; ZD_ENTRY=bing; BCLID=10732444505953897356; BCLID_BFESS=10732444505953897356; BDSFRCVID=lIuOJexroG33m7OJVAsYJTXpvQpWxY5TDYrELPfiaimDVu-VdK8uEG0Pts1-dEu-S2EwogKK0eOTHk8F_2uxOjjg8UtVJeC6EG0Ptf8g0M5; BDSFRCVID_BFESS=lIuOJexroG33m7OJVAsYJTXpvQpWxY5TDYrELPfiaimDVu-VdK8uEG0Pts1-dEu-S2EwogKK0eOTHk8F_2uxOjjg8UtVJeC6EG0Ptf8g0M5; H_BDCLCKID_SF=tRAOoC8-fIvDqTrP-trf5DCShUFsXjctB2Q-XPoO3K85DD3OyljNKpFp3b7j-qojQ5bk_xbgy4op8P3y0bb2DUA1y4vpKUPObmTxoUJ2-KDVeh5Gqq-KQJ-ebPRiJPQ9QgbWVpQ7tt5W8ncFbT7l5hKpbt-q0x-jLTnhVn0MBCK0hD89Dj-Ke5PthxO-hI6aKC5bL6rJabC3DhTGXU6q2bDeQN-e26b9WHcCQMTFQUn2fxjx3n7Zjq0vWq54WpOh2C60WlbCb664ehk95fonDh83KNLLKUQtHGAHK43O5hvvhb5O3M7OeMKmDloOW-TB5bbPLUQF5l8-sq0x0bOte-bQXH_E5bj2qRAJ_DDb3J; H_BDCLCKID_SF_BFESS=tRAOoC8-fIvDqTrP-trf5DCShUFsXjctB2Q-XPoO3K85DD3OyljNKpFp3b7j-qojQ5bk_xbgy4op8P3y0bb2DUA1y4vpKUPObmTxoUJ2-KDVeh5Gqq-KQJ-ebPRiJPQ9QgbWVpQ7tt5W8ncFbT7l5hKpbt-q0x-jLTnhVn0MBCK0hD89Dj-Ke5PthxO-hI6aKC5bL6rJabC3DhTGXU6q2bDeQN-e26b9WHcCQMTFQUn2fxjx3n7Zjq0vWq54WpOh2C60WlbCb664ehk95fonDh83KNLLKUQtHGAHK43O5hvvhb5O3M7OeMKmDloOW-TB5bbPLUQF5l8-sq0x0bOte-bQXH_E5bj2qRAJ_DDb3J; BDRCVFR[feWj1Vr5u3D]=I67x6TjHwwYf0; H_PS_PSSID=61027_61219_61238_60853_61362_61367_61390_61393_61422; PSINO=5; delPer=0; ZFY=yv9GDCzgJpoXNG5p2lUBb7cU0q1L:BuourCC4io748G8:C; BIDUPSID=016732EAAC60A2F9549AE8FF4A670FD8; PSTM=1733596883; HMACCOUNT=9A72F82981CECBA0; __bid_n=19367a1f6f78432ccc670d; BAIDU_WISE_UID=wapp_1728139968415_135; BAIDUID=016732EAAC60A2F9549AE8FF4A670FD8:FG=1'

    # Cookie = input("请输入你的Cookie，若错误则无法运行:")
    baidu_index = baidu_crawler(Cookie,keys,startDate,endDate)

    # 提示用户输入关键词
    words = input("请输入一串关键词(空格隔开): ")
    words = words.split()
    if words != []:
        keys = words

    # 定义直辖市/自治区列表
    direct_cities = ["北京", "上海", "天津", "重庆", "澳门", "香港"]
    
    cities = []

    while True:
        # 提示用户输入地域名称
        loc = input("请输入省/直辖市/自治区名称(如：北京/广东，输入all表示全部地区数据,输入 n 表示全国范围, 输入 q 退出)")
        if loc == "all" or loc in direct_cities:
            cities.append(loc)
        elif loc == "n":
            cities.append("全国")
        elif loc in p2c:
            city = input("请输入其中的城市 (例如: 广东-珠海，输入all表示省份数据)")
            if city in p2c.get(loc, []):
                loc = loc + "-" + city
                cities.append(loc)
            elif city == "all":
                cities.append(loc)
            else:
                print(f"{loc}没有此城市")
        elif loc == "q":
            break
        else:
            print("不存在此省市")
    if cities == []:
        # 默认城市
        cities = [
            "全国",
            "北京", "上海", "广东-广州", "广东-深圳",
            "安徽-合肥", "福建-福州", "福建-厦门", "福建-泉州",
            "甘肃-兰州", "广西-南宁", "贵州-贵阳", "黑龙江-哈尔滨",
            "河北-石家庄", "河南-郑州", "湖北-武汉", "湖南-长沙", 
            "江苏-南京", "江苏-苏州", "江苏-常州", "江苏-南通", 
            "江苏-无锡", "江苏-徐州", "江西-南昌", "吉林-长春", 
            "辽宁-沈阳", "辽宁-大连", "内蒙古-呼和浩特", "宁夏-银川", 
            "山东-济南", "山东-青岛", "山西-太原", "陕西-西安",
            "四川-成都",
            "天津", "新疆-乌鲁木齐", "云南-昆明", "浙江-杭州", 
            "浙江-宁波", "浙江-金华", "浙江-温州", "浙江-嘉兴", 
            "广东-佛山", "广东-东莞", "广东-惠州", "西藏-拉萨"
        ]
    baidu_index.crawler(cities)